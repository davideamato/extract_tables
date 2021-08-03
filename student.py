'''
    Contains the objects for table extraction 
'''

from collections import Counter
from utils import (InputError, desired_tables, detail_string,
                   escape_backslash_r, is_abs_path, math_mapping, physics_mapping,
                   fm_mapping, qualifications_with_overall_score, valid_exams,
                   ib_permutations)

from pandas import isna
import os


class GradeEntry:
    '''
        Class to store the grades
    '''

    def __init__(self, qualification, subject, grade, is_predicted, year, is_exam_result):
        if type(grade) is str:
            self.grade = escape_backslash_r(grade)
        else:
            self.grade = grade
        self.subject = escape_backslash_r(subject)
        self.qualification = escape_backslash_r(qualification)
        self.is_predicted = is_predicted
        self.is_exam_result = is_exam_result
        self.year = year

        # self.grade_info = [self.grade, self.subject, self.qualification, self.is_predicted, self.is_exam_result, self.year]
        self.grade_info = [self.grade, self.subject]

    # def __iter__(self):
    #     return self

    # def __next__(self):
    #     if self.index == 2:
    #         raise StopIteration

    #     self.index += 1

    #     return self.grade_info[self.index]

    def __repr__(self):
        return r"Qualification: {} Subject: {} Grade: {} Year: {} Predicted: {} Exam Result: {}".format(
            self.qualification, self.subject, self.grade, self.year, self.is_predicted, self.is_exam_result)

    def __str__(self):
        return r"Qualification: {} Subject: {} Grade: {} Year: {} Predicted: {} Exam Result: {}".format(
            self.qualification, self.subject, self.grade, self.year, self.is_predicted, self.is_exam_result)


class ExtractedStudents:
    '''
        Class that stores all the students and co-ordinates the output
    '''

    def __init__(self, applicant_ids, internal_mapping):
        self.student_ids = applicant_ids

        self.num_students = len(applicant_ids)

        self.all_students = [None]*self.num_students

        self.internal_mapping = internal_mapping

        self.index = -1

    def __iter__(self):
        return self

    def __next__(self):
        if self.index == self.num_students-1:
            raise StopIteration

        self.index += 1

        return self.all_students[self.index]

    def add_student_sequentially(self, new_student, counter):
        if new_student.ucas_id == self.student_ids[counter]:
            self.all_students[counter] = new_student
        else:
            raise RuntimeError("The order of adding students is incorrect \n"
                               "This should be done sequentially IDs in list should correspond")

    @staticmethod
    def populate_normal_header(ws):
        ws.cell(row=1, column=1, value="UCAS ID")
        ws.cell(row=1, column=2, value="Qualification Type")

        for i in range(0, 8, 2):
            ws.cell(row=1, column=3 + i, value="Subject")
            ws.cell(row=1, column=4 + i, value="Grade")

        return ws

    def populate_worksheet(self, desired_data, ws):
        if desired_data not in self.all_students[0].which_grades.keys():
            raise InputError(False, "Key given not found in dictionary")

        ws = self.populate_normal_header(ws)

        lb = 2

        for row_counter, student in zip(range(lb, self.num_students+lb), self.all_students):

            ws.cell(row=row_counter, column=1,
                    value="{}".format(student.ucas_id))

            if student.which_grades[desired_data]:
                ws.cell(row=row_counter, column=2, value="{}".format(
                    student.which_grades[desired_data][0].qualification))
                col_counter = 3

                for entry in student.which_grades[desired_data]:
                    ws.cell(row=row_counter, column=col_counter,
                            value="{}".format(entry.subject))
                    ws.cell(row=row_counter, column=col_counter +
                            1, value="{}".format(entry.grade))
                    col_counter += 2

        return ws

    @staticmethod
    def populate_master_header(ws):

        ws.cell(row=1, column=1, value="UCAS ID")
        ws.cell(row=1, column=2, value="Qualification")
        ws.cell(row=1, column=3, value="Issues Importing?")
        ws.cell(row=1, column=4, value="No. Subjects")
        ws.cell(row=1, column=5, value="Math Grade")
        ws.cell(row=1, column=6, value="Physics Grade")
        ws.cell(row=1, column=11, value="FM?")
        ws.cell(row=1, column=12, value="Overall Grade")

        for i in range(0, 4, 2):
            ws.cell(row=1, column=7 + i, value="Grade")
            ws.cell(row=1, column=8 + i, value="Subject")

        return ws

    @staticmethod
    def update_al_string(categorised_entries, input_string, is_fm):
        if is_fm:
            # If fm => min 3 subjects
            if categorised_entries['additional_subjects']:
                # If the list is not empty => 4 or more subjects
                return input_string + " (A*AAA)"
            else:
                # If the list is empty => 3 subjects
                return input_string + " (A*A*A)"
        else:
            if len(categorised_entries['additional_subjects']) > 1:
                return input_string + " (A*AAA)"
            else:
                # If the list is empty => 3 subjects
                return input_string + " (A*A*A)"

    def compile_for_master(self, ws):

        ws = self.populate_master_header(ws)

        lb = 2

        for row_counter, student in zip(range(lb, self.num_students+lb), self.all_students):
            # Fill in UCAS ID
            ws.cell(row=row_counter, column=1,
                    value="{}".format(student.ucas_id))

            # Categorise each entry into subjects
            categorised_entries = self.sort_into_subjects(student)

            # Identify if FM is presetn
            if categorised_entries["fm"]:
                is_fm = True
                ws.cell(row=row_counter, column=11, value="Yes")
            else:
                is_fm = False
                ws.cell(row=row_counter, column=11, value="No")

            # Identify and populate cell with the main qualification
            main_qualification = student.get_main_qualification()
            sanitised_string = self.internal_mapping[main_qualification]

            if "United Kingdom" in sanitised_string:
                uk_based = True
            else:
                uk_based = False

            if "A Levels" in sanitised_string:
                sanitised_string = self.update_al_string(
                    categorised_entries, sanitised_string, is_fm)

            # Fill in the qualification to the worksheet
            ws.cell(row=row_counter, column=2,
                    value="{}".format(sanitised_string))

            # Create log for issues
            any_issues = self.log_issues(categorised_entries, uk_based, ws, row_counter)

            # Get all unique qualifications
            qualification = student.unique_qualifications()

            # Identify if the qualification has an overall score
            # Intersection of qualification set and set of qualifications with overall score is not empty
            intersection_qualification = qualification & qualifications_with_overall_score()
            # If it does, find the overall score
            if intersection_qualification:

                # Is it IB?
                if intersection_qualification & ib_permutations():
                    if student.results_entries:
                        overall_grade = [
                            item.grade for item in student.results_entries if "IB Total points" in item.qualification]
                    elif student.completed_entries:
                        overall_grade = [
                            item.grade for item in student.completed_entries if "International Baccalaureate Diploma" in item.qualification]
                    elif student.predicted_entries:
                        overall_grade = [
                            item.grade for item in student.predicted_entries if "International Baccalaureate Diploma" in item.qualification]
                    else:
                        overall_grade = []
                else:
                    intersection_qualification = list(
                        intersection_qualification)
                    # Get all grades that aren't none
                    overall_grade = [item for qualification in intersection_qualification for item in student.get_grade_for_qualification(
                        qualification)]

                # If an overall grade exists, populate cell
                num_overall_grade = len(overall_grade)

                # If an overall grade exists
                if num_overall_grade > 1:
                    # Log multiple grades as an issue
                    if any_issues is not None:
                        any_issues.append(
                            "Multiple overall score. 1st selected.")
                    else:
                        # If any_issue is None => M&P grade not found. Skip rest of loop
                        ws.cell(row=row_counter, column=3,
                                value="M&P missing, need manual entry. Multiple overall score. 1st selected.")
                        ws.cell(row=row_counter, column=12,
                                value="{}".format(overall_grade[0]))
                        continue

                # Populate if list is not empty
                if num_overall_grade != 0:
                    ws.cell(row=row_counter, column=12,
                            value="{}".format(overall_grade[0]))

            # If M&P missing, clearly there is an issue. Skip the rest of loop
            if any_issues is None:
                ws.cell(row=row_counter, column=3,
                        value="M&P missing, need manual entry")
                continue

            # Populate subject and grades
            subjectCounter = 0
            map_subject_num_to_cols = {0: [5], 1: [6], 2: [7, 8], 3: [9, 10]}
            for subject_entries in categorised_entries.values():

                # If no entries, then go to next value
                if not subject_entries:
                    if subjectCounter > 1:
                        continue
                    else:
                        subjectCounter += 1
                        continue

                excel_col = map_subject_num_to_cols.get(subjectCounter)
                # If there is only one subject entry, then just populate
                if len(subject_entries) == 1:
                    for col, val in zip(excel_col, subject_entries[0].grade_info):
                        ws.cell(row=row_counter, column=col, value=val)

                # if it is at the 3rd subject, and not FM then iterate over all
                elif subjectCounter == 2 and not is_fm:
                    for entry in subject_entries[:2]:
                        excel_col = map_subject_num_to_cols.get(subjectCounter)
                        for col, val in zip(excel_col, entry.grade_info):
                            ws.cell(row=row_counter, column=col, value=val)
                        subjectCounter += 1

                # If there is FM and there are more than 4 subjects, populate with 1st
                elif subjectCounter == 3 and len(subject_entries) > 1:
                    for col, val in zip(excel_col, subject_entries[0].grade_info):
                        ws.cell(row=row_counter, column=col, value=val)

                # Doesn't fall into any of above cases => special case
                else:

                    exam_result_entries = [
                        entry for entry in subject_entries if entry.is_exam_result]
                    predicted_entries = [
                        entry for entry in subject_entries if entry.is_predicted]
                    excel_col = excel_col

                    map_counter_to_subject = {
                        0: "math", 1: "physics", 2: "3rd", 3: "4th"}
                    any_issues.append("For {}, selected ".format(
                        map_counter_to_subject[subjectCounter]))

                    # # REFACTOR
                    if len(exam_result_entries) == 1:
                        any_issues[-1] += "exam result"
                        selected = exam_result_entries[0]
                    elif len(exam_result_entries) > 1:
                        any_issues[-1] += "exam result"
                        selected = self.select_an_entry(
                            exam_result_entries, any_issues)
                    elif len(predicted_entries) == 1:
                        any_issues[-1] += "predicted grade"
                        selected = predicted_entries[0]
                    elif len(predicted_entries) > 1:
                        any_issues[-1] += "predicted grade"
                        selected = self.select_an_entry(
                            predicted_entries, any_issues)
                    else:
                        any_issues[-1] += "from all"
                        selected = self.select_an_entry(
                            subject_entries, any_issues)

                    for col, val in zip(excel_col, selected.grade_info):
                        ws.cell(row=row_counter, column=col, value=val)

                subjectCounter += 1

            # Compress list of strings into a single string
            any_issues = self.compress_log(any_issues)
            # Populate cell with string
            if any_issues is None:
                ws.cell(row=row_counter, column=3, value="")
            else:
                ws.cell(row=row_counter, column=3, value=any_issues)

        return ws

    @staticmethod
    def select_an_entry(lst_of_entries, any_issues):
        # ASSUMPTION: MOST RECENT GRADE = BEST GRADE

        years = [entry.year for entry in lst_of_entries if entry.year is not None]

        # If the list is empty then look for highest grade
        if not years:
            grades = [str(entry.grade)
                      for entry in lst_of_entries if entry.grade is not None]
            # If this list is not empty, no issues. Take the highest
            if grades:
                any_issues[-1] += "highest "
                return lst_of_entries[grades.index(max(grades))]
            else:
                any_issues[-1] = "No grade"
                return GradeEntry(None, None, "None", False, None, False)

        # Determine if all are unique and whether the lsit is empty
        if len(years) == len(set(years)) and years:
            # All unique years
            any_issues[-1] += "most recent "

            return lst_of_entries[years.index(max(years))]

        max_year = max(years)
        most_recent = [
            entry for entry in lst_of_entries if entry.year == max_year]
        if len(most_recent) > 1:
            # Max year is repeated

            grades = [str(entry.grade)
                      for entry in most_recent if entry.grade is not None]
            any_issues[-1] += "highest most recent "

            return lst_of_entries[grades.index(max(grades))]

    @staticmethod
    def compress_log(log):
        if len(log) > 1:
            return " ".join(log)
        elif len(log) == 1:
            return log[0]
        else:
            return None

    @staticmethod
    def log_issues(categorised_entries, uk_based, ws, rowCounter):
        convert_lst_to_bool = [not bool(cat_entry)
                               for cat_entry in categorised_entries.values()]
        # All lists in categorised entries are empty
        if all(convert_lst_to_bool):
            return "No valid qual & subjects."

        if all(convert_lst_to_bool[:3]):
            return None

        log = []

        for subject, entries in categorised_entries.items():
            entries_length = len(entries)
            if entries_length > 1 and subject != "additional_subjects":
                log.append("Multiple {} qual.".format(subject))
            elif entries_length > 2 and subject == "additional_subjects":
                if not uk_based:
                    # log.append("Multiple Subjects.")
                    ws.cell(row=rowCounter, column=4, value="Multiple")
                else:
                    # log.append(">4 Subjects.")
                    ws.cell(row=rowCounter, column=4, value=">4")
            elif entries_length == 0 and subject != "fm" and subject != "additional_subjects":
                log.append("{} missing.".format(subject))

        return log

    @staticmethod
    def sort_into_subjects(student):

        categorised_entries = {
            "math": [],
            "physics": [],
            "fm": [],
            "additional_subjects": [],
        }

        # Grade type => results/predicted/completed
        for grade_entries in student.which_grades.values():
            # List of entries are not empty
            if grade_entries:
                for entry in grade_entries:
                    if entry.subject in math_mapping().get(entry.qualification, set()):
                        categorised_entries["math"].append(entry)
                    elif entry.subject in physics_mapping().get(entry.qualification, set()):
                        categorised_entries["physics"].append(entry)
                    elif entry.subject in fm_mapping().get(entry.qualification, set()):
                        categorised_entries["fm"].append(entry)
                    else:
                        categorised_entries["additional_subjects"].append(
                            entry)

        return categorised_entries

    def write_to_excel(self, input_abs_path):
        is_abs_path(input_abs_path)

        from openpyxl import Workbook

        wb = Workbook()

        # Create and populate workshets
        completed = wb.create_sheet("Completed Qualifications")
        completed = self.populate_worksheet("completed", completed)

        predicted = wb.create_sheet("Predicted Grades")
        predicted = self.populate_worksheet("predicted", predicted)

        exam_results = wb.create_sheet("Exam Results")
        exam_results = self.populate_worksheet("results", exam_results)

        compiled_single = wb.create_sheet("Compiled", 0)
        compiled_single = self.compile_for_master(compiled_single)

        wb.save(os.path.join(input_abs_path, "output.xlsx"))


class StudentGrades:
    '''
        Class for a single pdf/student
    '''

    def __init__(self, id, extracted_tables, table_headers):
        self.ucas_id = id

        self.completed_qualifications = None
        self.uncompleted_qualifications = None
        self.exam_results = None

        target_tables = desired_tables()

        for header, tbl in zip(table_headers, extracted_tables):
            if header == target_tables[0]:
                # print("")
                # print("Completed Qualification")
                # print(tbl)
                self.completed_qualifications = tbl
            elif header == target_tables[1]:
                # print("")
                # print("Predicted Grades")
                # print(tbl)
                self.uncompleted_qualifications = tbl
            elif header == target_tables[2]:
                # print("")
                # print("Exam Results")
                # print(tbl)
                self.exam_results = tbl

        self.predicted_entries = []
        self.completed_entries = []
        self.results_entries = []

        self.predicted_grade_entries()
        self.examresult_entries()
        self.completed_grade_entries()

        self.which_grades = {
            "results": self.results_entries,
            "completed": self.completed_entries,
            "predicted": self.predicted_entries,
        }

    def __repr__(self):
        return "{} \n {} \n {}".format(self.completed_qualifications,
                                       self.uncompleted_qualifications,
                                       self.exam_results)

    def get_all_qualifications(self):
        return [item.qualification for grade_entries in self.which_grades.values(
        ) if grade_entries for item in grade_entries]

    def unique_qualifications(self):
        return set(self.get_all_qualifications())

    def get_main_qualification(self):
        qualifications = self.get_all_qualifications()
        if not qualifications:
            return ""

        if len(set(qualifications)) == 1:
            return qualifications[0]
        else:
            return Counter(qualifications).most_common(1)[0][0]

    def get_grade_for_qualification(self, target_qualification):
        for values in self.which_grades.values():
            if values:
                for item in values:
                    # It is the qualification we are looking for.
                    # The grade is not None AND year is not None (implies it is a module/detail entry)
                    if target_qualification in item.qualification and item.grade is not None and item.year is not None:
                        yield item.grade

    def is_detailed_entry(self, input_qualification, rowCounter):
        target = input_qualification['Date'][rowCounter]
        if type(target) is not str:
            return False

        if target not in detail_string():
            return False

        return True

    def handle_detailed_entry(self, input_qualification, rowCounter):

        if "Exam" in set(input_qualification.columns):
            qualification_identifier = "Exam"
        elif "Exam Level" in set(input_qualification.columns):
            qualification_identifier = "Exam Level"
        else:
            raise NotImplementedError

        if not isna(input_qualification[qualification_identifier][rowCounter-1]):
            qualification = input_qualification[qualification_identifier][rowCounter-1]
        else:
            qualification = None

        output = []
        all_module_details = input_qualification['Body'][rowCounter]

        # Ignores the first entry which would just be the date
        individual_modules = all_module_details.split("Title:")[1:]
        # print(individual_modules)
        for module in individual_modules:

            module_info = module.split("Date:")[0]

            if "Predicted Grade:" in module_info:
                grade = module_info.split("Predicted Grade:")[0]
            elif "Grade:" in module_info:
                grade = module_info.split("Grade:")[0]
            elif "Value:" in module_info:
                grade = module_info.split("Value:")[0]
            else:
                grade = None

            entry = GradeEntry(
                qualification,
                module_info,
                grade,
                True,
                None,
                False,
            )
            output.append(entry)

        return output

    def completed_grade_entries(self):
        if self.completed_qualifications is None:
            return None

        for row in self.completed_qualifications.index:

            if self.is_completed_qual_valid(row):

                entry = GradeEntry(
                    self.completed_qualifications['Exam'][row],
                    self.completed_qualifications['Subject'][row],
                    self.completed_qualifications['Grade'][row],
                    False,
                    self.completed_qualifications['Date'][row].split("-")[-1],
                    False,
                )

                self.completed_entries.append(entry)

            elif self.is_detailed_entry(self.completed_qualifications, row) and self.is_completed_qual_valid(row-1):

                detailed_entries = self.handle_detailed_entry(
                    self.completed_qualifications, row)
                self.completed_entries += detailed_entries

        return self.completed_entries

    def is_completed_qual_valid(self, row):
        if isna(self.completed_qualifications['Exam'][row]):
            return False

        if self.completed_qualifications['Exam'][row] in valid_exams():
            return True
        else:
            return False

    def examresult_entries(self):
        if self.exam_results is None:
            return None

        for row in self.exam_results.index:

            if self.is_examresult_valid(row):

                entry = GradeEntry(
                    self.exam_results['Exam Level'][row],
                    self.exam_results['Subject'][row],
                    self.exam_results['Grade'][row],
                    False,
                    self.exam_results['Date'][row].split("-")[-1],
                    True,
                )

                self.results_entries.append(entry)

            elif self.is_detailed_entry(self.exam_results, row):

                detailed_entries = self.handle_detailed_entry(
                    self.exam_results, row)
                self.results_entries += detailed_entries

        return self.results_entries

    def is_examresult_valid(self, row):
        if isna(self.exam_results['Exam Level'][row]):
            return False

        if self.exam_results['Exam Level'][row] in valid_exams():
            return True
        else:
            return False

    def predicted_grade_entries(self):
        if self.uncompleted_qualifications is None:
            return None

        for row in self.uncompleted_qualifications.index:

            is_pred_grade = isna(
                self.uncompleted_qualifications['Predicted\rGrade'][row])
            is_grade = isna(self.uncompleted_qualifications['Grade'][row])

            if is_pred_grade ^ is_grade:

                if is_pred_grade:
                    valid_grade = self.uncompleted_qualifications['Grade'][row]
                else:
                    valid_grade = self.uncompleted_qualifications['Predicted\rGrade'][row]

                if isna(self.uncompleted_qualifications['Exam'][row]):
                    qualification = self.uncompleted_qualifications['Body'][row]
                else:
                    qualification = self.uncompleted_qualifications['Exam'][row]

                entry = GradeEntry(
                    qualification,
                    self.uncompleted_qualifications['Subject'][row],
                    valid_grade,
                    True,
                    self.uncompleted_qualifications['Date'][row].split(
                        "-")[-1],
                    False,
                )
                self.predicted_entries.append(entry)

            elif (not is_pred_grade) & (not is_grade):

                if "Unnamed" in self.uncompleted_qualifications['Grade'][row]:
                    valid_grade = self.uncompleted_qualifications['Predicted\rGrade'][row]
                else:
                    valid_grade = self.uncompleted_qualifications['Grade'][row]

                if isna(self.uncompleted_qualifications['Exam'][row]):
                    qualification = self.uncompleted_qualifications['Body'][row]
                else:
                    qualification = self.uncompleted_qualifications['Exam'][row]

                entry = GradeEntry(
                    qualification,
                    self.uncompleted_qualifications['Subject'][row],
                    valid_grade,
                    True,
                    self.uncompleted_qualifications['Date'][row].split(
                        "-")[-1],
                    False
                )
                self.predicted_entries.append(entry)

            elif type(self.uncompleted_qualifications['Date'][row]) is str:
                if is_pred_grade & is_grade and self.uncompleted_qualifications['Date'][row] in detail_string():
                    all_module_details = self.uncompleted_qualifications[
                        'Body'][row]

                    if isna(self.uncompleted_qualifications['Exam'][row-1]):
                        qualification = self.uncompleted_qualifications['Body'][row-1]
                    else:
                        qualification = self.uncompleted_qualifications['Exam'][row-1]

                    # Ignores the first entry which would just be the date
                    individual_modules = all_module_details.split("Title:")[1:]
                    # print(individual_modules)
                    for module in individual_modules:
                        module_info = module.split("Date:")[0]

                        if "Predicted Grade:" in module_info:
                            split = module_info.split("Predicted Grade:")
                            subject = split[0]
                            grade = split[1]
                        elif "Grade:" in module_info:
                            split = module_info.split("Grade:")
                            subject = split[0]
                            grade = split[1]
                        elif "Value:" in module_info:
                            split = module_info.split("Value:")
                            subject = split[0]
                            grade = split[1]
                        else:
                            subject = module_info
                            grade = None

                        entry = GradeEntry(
                            qualification,
                            subject,
                            grade,
                            True,
                            None,
                            False,
                        )
                        self.predicted_entries.append(entry)

        return self.predicted_entries
