import os
import logging
import csv
import shutil

from time import localtime, strftime
from copy import deepcopy
from random import randint

import tabula
import numpy as np
from pandas import read_excel, Series

import settings

from pdf_strings import detail_string


class InputError(Exception):
    """Exception raised for errors in the input.

    Attributes:
        expression -- input expression in which the error occurred
        message -- explanation of the error
    """

    def __init__(self, expression, message):
        self.expression = expression
        self.message = message


def initialise_logger():

    if os.path.exists(settings.path_to_log):
        os.remove(settings.path_to_log)

    logging.basicConfig(
        format="%(asctime)s %(levelname)s:%(message)s",
        filename=settings.path_to_log,
        datefmt="%m/%d/%Y %I:%M:%S %p",
        level=logging.INFO,
    )
    logging.info("Start")


def get_internal_mapping(path_to_file, sheet_name):
    if not path_to_file.endswith(".xlsx"):
        logging.error("Mapping file not in xlsx format")
        raise InputError(
            'not path_to_file.endswith(".xlsx")', "Input file must be in xlsx format"
        )

    input_file = path_to_file

    from openpyxl import load_workbook

    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb[sheet_name]

    # Maps values in columns after 2nd to value in 1st column
    output_dict = dict()

    for row in ws.rows:
        val = None
        for cell in row:
            if val is None:
                # Value in first column is the desired string
                val = cell.value
            elif cell.value is not None:
                # Other values is what we have
                output_dict[''.join(e for e in str(cell.value) if e.isprintable() and not e.isspace())] = val

    logging.info("Mapping file loaded")

    return output_dict


def get_subject_mapping(path_to_file, sheet_name):
    if not path_to_file.endswith(".xlsx"):
        logging.error("Mapping file not in xlsx format")
        raise InputError(
            'not path_to_file.endswith(".xlsx")', "Input file must be in xlsx format"
        )

    input_file = path_to_file

    from openpyxl import load_workbook

    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb[sheet_name]

    # Maps values in columns after 2nd to value in 1st column
    output_dict = dict()

    for row in ws.rows:
        qual = None
        for cell in row:
            if qual is None:
                # Value in first column is the qualification
                qual = cell.value
                output_dict[qual] = set()
            elif cell.value is not None:
                # Other values is subject names
                output_dict[qual].add(cell.value)

    return output_dict


def is_file_valid(file):
    if file.endswith(".pdf") and "unicode" in file:
        return True
    else:
        logging.info(f"{file} is not valid. Unicode or .pdf not in file name")
        return False


def is_abs_path(input_path):
    if not os.path.isabs(input_path):
        logging.error("Absolut path to file not provided")
        raise InputError(
            "not os.path.isabs(input_path)", "Path provided is not absolute"
        )

    return True


def check_output_dirs_exist():
    marker_names = list(settings.allocation_details.keys())

    for name in marker_names:
        folder_name = name + str(settings.batch_number)
        marker_path = os.path.join(settings.output_path, folder_name)
        if not os.path.exists(marker_path):
            os.makedirs(marker_path)


def check_target_id_file_settings():
    if settings.is_id_file_banner:
        if settings.which_column is None:
            raise InputError(
                "settings.is_id_file_banner and settings.which_column is None",
                "If ID file is banner, which column must be specified",
            )
        elif settings.is_banner_cumulative is None:
            raise InputError(
                "settings.is_id_file_banner and settings.is_banner_cumulative is None",
                "If ID file is banner, this must be True or False (a boolean)",
            )
    else:
        if settings.is_banner_cumulative:
            raise InputError(
                "settings.is_banner_cumulative and settings.is_id_file_banner",
                "If ID is NOT banner, is_banner_cumulative must be False",
            )
        elif settings.which_column is not None:
            raise InputError(
                "not settings.is_id_file_banner and settings.which_column is not None",
                "If ID is NOT banner, which_column must be None",
            )


def is_database_path_valid():
    if settings.path_to_database_of_extracted_pdfs is None:
        if settings.is_banner_cumulative:
            raise InputError(
                True,
                (
                    "Database not provided BUT target file is cumulative."
                    + "\nPlease provide a path to database"
                ),
            )
        else:
            return False

    _, ext = os.path.splitext(settings.database_name)

    if ext != ".csv":
        raise InputError(True, "Database file MUST be a .csv")

    return True


def handle_banner_and_database_permutations(ids_from_database):
    msg = None
    if settings.is_banner_cumulative and ids_from_database is None:
        msg = "WARNING: No database values found but banner is cumulative"
        if settings.batch_number != 1:
            msg += f" and batch number ({settings.batch_number}) != 1 "
            raise InputError("settings.batch_number != 1", msg)
        else:
            msg += (
                "\n\t Batch number is 1. Will continue on assumption of no previous IDs"
            )
            msg += "\n\t Please terminate if this is incorrect"
    elif not settings.is_banner_cumulative and ids_from_database is not None:
        msg = (
            "WARNING: Database values found but banner is not cumulative"
            + "\nPlease set database path to None if not needed"
        )

    if msg is not None:
        print(msg)
        logging.warning(msg)


def get_ids_from_database():
    if settings.is_id_file_banner:
        if not is_database_path_valid():
            return None

        ids_from_database = get_previous_ids(
            settings.path_to_database_of_extracted_pdfs
        )
        # print(ids_from_database)
        handle_banner_and_database_permutations(ids_from_database)
    else:
        ids_from_database = None

    return ids_from_database


def get_data_from_target_file():
    if settings.is_id_file_banner:
        data_from_sheet = read_excel(
            settings.path_to_target_file,
            engine="openpyxl",
            usecols=settings.which_column,
            dtype=int,
        )
        # print(data_from_sheet)
    else:
        data_from_sheet = read_excel(
            settings.path_to_target_file, engine="openpyxl", dtype=int, header=None
        )
    return data_from_sheet.values.flatten().tolist()


def check_ids_correspond(ids_from_pdf_folder):

    check_target_id_file_settings()

    ids_from_target_file = get_data_from_target_file()

    order_preserved_target_ids = deepcopy(ids_from_target_file)

    # Enforce same type - both Integers
    ids_from_pdf_folder = [int(item) for item in ids_from_pdf_folder]

    # Convert to set to allow for obtain desired set
    ids_from_pdf_folder = set(ids_from_pdf_folder)
    ids_from_target_file = set(ids_from_target_file)
    ids_from_database = get_ids_from_database()

    if ids_from_database is not None:
        ids_from_database = set(ids_from_database)

        # print(ids_from_target_file)
        # print(ids_from_pdf_folder)
        # print(ids_from_database)

        if ids_from_target_file < ids_from_database:
            not_in_target = ids_from_database - ids_from_target_file
            # not_in_target = list(not_in_target)
            not_in_target = ", ".join([str(item) for item in list(not_in_target)])
            msg = (
                "Database IDs is a superset of target IDs"
                + f"Following IDs in database but not in target ids file: {not_in_target}"
            )
            print(msg)
            logging.error(msg)
            raise InputError(
                "not ids_from_target_file.issuperset(ids_from_database)",
                "Target ids file is not a super set of database IDs",
            )
        elif ids_from_target_file < ids_from_pdf_folder:
            not_in_target = ids_from_pdf_folder - ids_from_target_file
            not_in_target = ", ".join([str(item) for item in list(not_in_target)])
            msg = (
                "PDF IDs is a superset of target IDs"
                + f"Following IDs in PDFs but not in target ids file: {not_in_target}"
            )
            print(msg)
            logging.warning(msg)
        elif ids_from_pdf_folder - ids_from_target_file:
            not_in_target = ids_from_pdf_folder - ids_from_target_file
            not_in_target = ", ".join([str(item) for item in list(not_in_target)])
            msg = f"Following IDs in PDFs but not in target ids file: {not_in_target}"
            print(msg)
            logging.warning(msg)

        # New IDs are defined as IDs in target file but not in database
        new_ids = ids_from_target_file - ids_from_database

        if not new_ids:
            raise InputError("not new_ids", "TERMINATE: No new IDs")

        if ids_from_pdf_folder.issuperset(new_ids):
            # return list(new_ids)
            return [target_id for target_id in order_preserved_target_ids if target_id in new_ids]
        else:
            missing_ids = new_ids - ids_from_pdf_folder
            missing_ids = ", ".join([str(item) for item in list(missing_ids)])
            msg = f"Following ID(s) are new but PDF not found: {missing_ids}"
            logging.error(msg)
            raise InputError("not ids_from_pdf_folder.issuperset(new_ids)", msg)

    else:

        intersection = ids_from_target_file & ids_from_pdf_folder

        # Intersection lap is empty => No overlap
        if ids_from_target_file.isdisjoint(ids_from_pdf_folder):
            # if not intersection:
            logging.error(
                f"Mismatch in IDs."
                f"IDs in {settings.path_to_pdfs_to_extract} folder don't "
                f"match IDs in {settings.target_ucas_id_file} file"
            )
            raise InputError(
                "not ids_from_target_file.isdisjoint(ids_from_pdf_folder)",
                "IDs from PDF folder does not match IDs to extract in Excel file",
            )

        if ids_from_pdf_folder > ids_from_target_file or (
            ids_from_pdf_folder - ids_from_target_file
        ):
            not_in_target = ids_from_pdf_folder - ids_from_target_file
            not_in_target = ", ".join([str(item) for item in list(not_in_target)])
            msg = (
                "WARNING!! \n"
                + f"Following ID(s) in PDF folder but not in target file: {not_in_target} \n"
                + "ONLY processsing IDs in target file \n"
                + "IDs listed above will be IGNORED"
            )
            print(msg)
            logging.warning(msg)

            return order_preserved_target_ids
        elif ids_from_target_file == intersection:
            return order_preserved_target_ids
        else:
            file_not_found = ids_from_target_file - intersection
            file_not_found = ", ".join([str(item) for item in list(file_not_found)])
            msg = f"Following ID(s) in target file but PDF not found: {file_not_found}"
            logging.error(msg)
            raise InputError("ids_from_target_file != intersection", msg)

def remove_extra_pdfs(target_ids, pdf_paths, pdf_ids):

    not_in_target = list(set(pdf_ids) - set(target_ids))
    extra_locs = [np.argwhere(np.asarray(pdf_ids) == id) for id in not_in_target]

    filtered_pdf_ids = [pdf_ids[i] for i in range(0, len(pdf_ids)) if i not in extra_locs]
    filtered_pdf_paths = [pdf_paths[i] for i in range(0, len(pdf_paths)) if i not in extra_locs]

    return (filtered_pdf_paths, filtered_pdf_ids) 

def order_pdfs_to_target_id_input(all_pdf_paths, ids_from_all_pdfs):

    # Perform check to see if IDs from PDFs and target IDs correspond
    target_ids = check_ids_correspond(ids_from_all_pdfs)

    # Convert to numpy array to get argwhere to work
    if isinstance(target_ids, list):
        target_ids = np.asarray(target_ids)
    else:
        target_ids = np.asarray(list(target_ids))

    if not isinstance(ids_from_all_pdfs, list):
        ids_from_all_pdfs = list(ids_from_all_pdfs)

    # Enforced type being integer for comparison
    ids_from_all_pdfs = [int(item) for item in ids_from_all_pdfs]
    # Remove extra pdfs from list
    all_pdf_paths, ids_from_all_pdfs = remove_extra_pdfs(target_ids, all_pdf_paths, ids_from_all_pdfs)

    # Location of ID from pdfs in target ids list
    id_locs = [
        np.argwhere(target_ids == current_id).item()
        for current_id in ids_from_all_pdfs
        if current_id in target_ids
    ]

    # Sort list based on id_locs, then extract the paths from it
    sorted_pdf_paths = [
        path
        for _, path in sorted(zip(id_locs, all_pdf_paths), key=lambda pair: pair[0])
    ]

    target_ids = [str(item) for item in target_ids.tolist()]

    return sorted_pdf_paths, target_ids


def copy_file(path_to_file, extracted_students_instance, id_num):
    marker_details = extracted_students_instance.student_to_marker_mapping.get(id_num)

    if marker_details is not None:
        marker_name, numbering = marker_details
    else:
        print(f"ID: {id_num} not found in assignment of student to marker")
        print("\t Continuing by randomly assigning on the fly")
        rand_index = randint(0, len(settings.allocation_details))
        marker_name = list(settings.allocation_details.keys())[rand_index]
        numbering = 0
        print(f"\t Marker {marker_name} selected and numbering set to 0")

    original_filename = os.path.basename(path_to_file)

    new_filename = str(numbering) + "_" + original_filename
    path_to_new_dir = os.path.join(settings.output_path, marker_name, new_filename)

    shutil.copy(path_to_file, path_to_new_dir)


def copy_pdfs_to_pool(all_pdf_paths):
    output_path = settings.path_to_pdf_pool
    for file in all_pdf_paths:
        shutil.copy(file, output_path)


def get_files_and_ids(abs_path):

    # Test if path is absolute
    is_abs_path(abs_path)

    logging.info("Assembling list of files to analyse...")

    # Remove duplicates in files
    # BUT doesn't address repeated IDs
    lst_of_paths = list(
        set(
            [
                os.path.join(abs_path, file)
                for file in os.listdir(abs_path)
                if is_file_valid(file)
            ]
        )
    )

    # Extract IDs from file names
    lst_of_ids = [
        os.path.basename(file).split(settings.pdf_filename_split_delimeter)[
            settings.pdf_filename_split_index
        ]
        for file in lst_of_paths
    ]

    num_files = len(lst_of_paths)
    logging.info(f"Total of {num_files} files")
    print(f"Total of {num_files} files in {settings.path_to_pdfs_to_extract}")

    # If all are unique, then no issues
    if len(lst_of_ids) == len(set(lst_of_ids)):
        logging.info("No duplicate files")
        print("No duplicate files")
        return lst_of_paths, lst_of_ids

    # If there are repetitions, remove it from the lsit of IDs and paths
    counter = 1
    num_repetitions = 0
    for unique_id in lst_of_ids[1:]:
        if unique_id in lst_of_ids[:counter]:
            logging.info(
                f"Duplicate file present for {unique_id}, file {lst_of_paths[counter]} removed"
            )
            lst_of_ids.pop(counter)
            lst_of_paths.pop(counter)
            num_repetitions += 1
        else:
            counter += 1

    logging.info(f"Total of {num_repetitions} repeated files removed")
    print(f"Total of {num_repetitions} repeated files excluded")
    print("Check log for details")

    return lst_of_paths, lst_of_ids


def check_batch_num_against_database(past_batch_nums, is_same_then_terminate=False):
    # Check and get user input on whether to continue based on batch number
    prev_max_batch_num = max(past_batch_nums)
    print(
        f"Current batch number: {settings.batch_number}\t"
        + f"Largest previous batch number: {prev_max_batch_num}"
    )
    if prev_max_batch_num > settings.batch_number:
        raise InputError(
            "prev_max_batch_num > settings.batch_number",
            f"Current batch number ({settings.batch_number}) is less than largest previous batch number ({prev_max_batch_num})",
        )
    elif prev_max_batch_num == settings.batch_number:
        if is_same_then_terminate:
            raise InputError(
                "prev_max_batch_num == settings.batch_number",
                "Please increment batch number by 1",
            )
        else:
            get_batch_continue_input(prev_max_batch_num)


def get_batch_continue_input(prev_max_batch_num):
    print(
        f"Current batch number is {settings.batch_number} and is the same as max. previous batch number ({prev_max_batch_num}) "
    )
    print("Is this correct? yes/no")
    is_correct = input()
    while is_correct not in {"yes", "no"}:
        is_correct = input("Please enter 'yes' or 'no' ")

    if is_correct == "no":
        raise Exception


def read_database_file(database_path):
    database_ids = []
    past_batch_nums = set()

    with open(database_path, "r") as database_file:
        database_reader = csv.DictReader(database_file, delimiter=",")

        # Put past ids into list, other information is for human
        for row in database_reader:
            row_id = row.get(
                settings.database_headers[settings.database_header_id_num_index]
            )
            if row_id is not None:
                database_ids.append(int(row_id))
            else:
                raise Exception("Database file has been corrupted")

            row_batch_num = row.get(
                settings.database_headers[settings.database_header_batch_index]
            )
            if row_batch_num is not None:
                past_batch_nums.add(int(row_batch_num))
            else:
                raise Exception("Database file has been corrupted")

        check_batch_num_against_database(
            past_batch_nums,
            is_same_then_terminate=settings.terminate_if_batch_num_repeated,
        )

    return database_ids


def get_previous_ids(database_path):
    if os.path.exists(database_path):
        database_ids = read_database_file(database_path)
        # Return None if list is empty
        if database_ids:
            return database_ids
        else:
            return None
    else:
        return None


def get_current_time():
    return strftime("%Y-%m-%d %H:%M", localtime())


def update_previous_id_database(database_path, new_ids):
    if os.path.exists(database_path):
        is_existing_file = True
        open_mode = "a"
    else:
        is_existing_file = False
        open_mode = "w"

    timestamp = strftime("%Y-%m-%d %H:%M", localtime())

    with open(database_path, open_mode) as database_file:
        database_writer = csv.DictWriter(
            database_file, fieldnames=settings.database_headers, delimiter=","
        )

        if not is_existing_file:
            database_writer.writeheader()

        for id_num in new_ids:
            database_writer.writerow(
                {
                    settings.database_headers[
                        settings.database_header_id_num_index
                    ]: id_num,
                    settings.database_headers[
                        settings.database_header_batch_index
                    ]: settings.batch_number,
                    settings.database_headers[
                        settings.database_header_timestamp_index
                    ]: timestamp,
                }
            )


def check_broken_table(current_page_number, filename, current_table):
    """
    Determines if a table continues onto the next page
    If it does, return a the data in a form that can be appended to the original table
    """

    # Extract tables from next page
    tables = tabula.read_pdf(
        filename,
        pages=str(current_page_number + 1),
        lattice=True,
        guess=True,
        pandas_options={"header": 0},
    )

    if not tables:
        return None

    # Table at the top of the page
    top_table = tables[0]
    top_table_header = top_table.columns

    current_table_header = current_table.columns
    current_table_length = len(current_table_header)

    if top_table.empty:
        table_length = len(top_table_header)
        if table_length == current_table_length:
            return Series(top_table_header.values, index=current_table_header)
        elif detail_string() in top_table_header:
            return move_data_out_of_header(
                top_table, current_table_header, table_length
            )
        else:
            return None

    elif len(top_table_header) == current_table_length:
        return move_data_out_of_header(
            top_table, current_table_header, current_table_length
        )
    else:
        return None


def move_data_out_of_header(top_table, cur_table_header, table_length):

    from pandas import DataFrame as pdDF

    # Moves header into next row
    top_table = top_table.reset_index().T.reset_index().T

    # But, I have a new cloumn, so delete
    del top_table[0]

    # Create a new data frame instead of renaming columns as that creates issues
    # Only select table headers that are relevant for table at top of the page
    return pdDF(top_table.values, columns=cur_table_header[:table_length])


def fix_broken_table(current_page_number, current_table, filename):

    continued_values = check_broken_table(current_page_number, filename, current_table)

    if continued_values is not None:
        # add new row to end of DataFrame
        updated_table = current_table.append(
            continued_values, ignore_index=True, sort=False
        )

        # print(updated_table)
        return updated_table
    else:
        return current_table


def escape_backslash_r(input_string):
    if input_string is not None:
        return (
            input_string.encode("unicode-escape").decode().replace("\\r", " ").strip()
        )
