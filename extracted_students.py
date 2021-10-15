import re
import os

from random import randint

from pdf_strings import (
    math_mapping,
    physics_mapping,
    fm_mapping,
    qualifications_with_overall_score,
    ib_permutations,
)

from utils import (
    InputError,
    is_abs_path,
)

import settings
from grade_entry import GradeEntry


class ExtractedStudents:
    """
    Class that stores all the students and co-ordinates the output
    """

    def __init__(self, applicant_ids, internal_mapping):
        self.student_ids = applicant_ids

        self.num_students = len(applicant_ids)

        self.all_students = [None] * self.num_students

        self.internal_mapping = internal_mapping

        self.marker_allocation = self.assign_students_to_marker()
        # print(self.marker_allocation)

        self.student_to_marker_mapping = dict()
        self.map_student_to_marker()
        # print(self.student_to_marker_mapping)

        self.index = -1

    def __iter__(self):
        return self

    def __next__(self):
        if self.index == self.num_students - 1:
            raise StopIteration

        self.index += 1

        return self.all_students[self.index]

    def add_student_sequentially(self, new_student, counter):
        if new_student.unique_id == self.student_ids[counter]:
            self.all_students[counter] = new_student
        else:
            raise RuntimeError(
                "The order of adding students is incorrect \n"
                "This should be done sequentially IDs in list should correspond"
            )

    def assign_students_to_marker(self):
        # Get details from settings
        allocation_details = settings.allocation_details

        # Get ratios from details
        # ratio = [val for val in allocation_details.values()]
        ratio = list(allocation_details.values())

        # Total number of parts
        num_parts = sum(ratio)
        num_markers = len(allocation_details)

        # Num students per part
        # Integer division
        students_per_part = self.num_students // num_parts

        # Leftover
        remainder = self.num_students % num_parts

        # Allocate based of num parts per marker
        num_allocated = [parts * students_per_part for parts in ratio]

        # Randomly add remainder to markers
        for _ in range(remainder):
            index = randint(0, num_markers - 1)
            num_allocated[index] = num_allocated[index] + 1

        allocation = dict()
        start_index = 0
        for initials, num_students in zip(allocation_details.keys(), num_allocated):
            # Create the name of folder
            name = initials + str(settings.batch_number)

            # Allocate by slicing list
            allocation[name] = self.student_ids[
                start_index : start_index + num_students
            ]

            # Write text file containing IDs to folder
            path_to_marker_pdfs = os.path.join(
                settings.output_path, name, settings.ids_in_folder_file
            )
            with open(path_to_marker_pdfs, "w") as file:
                file.write(" ,".join(allocation[name]))

            start_index += num_students

        return allocation

    def map_student_to_marker(self):
        for marker, assigned_ids in self.marker_allocation.items():
            counter = 0
            for student in assigned_ids:
                self.student_to_marker_mapping[student] = [marker, counter]
                counter += 1

        # return self.student_to_marker_mapping

    @staticmethod
    def populate_normal_header(ws):
        ws.cell(row=1, column=1, value="UCAS ID")
        ws.cell(row=1, column=2, value="Qualification Type")

        for i in range(0, 8, 2):
            ws.cell(row=1, column=3 + i, value="Subject")
            ws.cell(row=1, column=4 + i, value="Grade")

        return ws

    def populate_worksheet(self, desired_data, ws):
        if desired_data not in self.all_students[0].which_grades.keys():
            raise InputError(False, "Key given not found in dictionary")

        ws = self.populate_normal_header(ws)

        lb = 2

        for row_counter, student in zip(
            range(lb, self.num_students + lb), self.all_students
        ):

            ws.cell(row=row_counter, column=1, value="{}".format(student.unique_id))

            target_data = student.which_grades.get(desired_data)

            if target_data is not None:
                if len(target_data) > 0:
                    ws.cell(
                        row=row_counter,
                        column=2,
                        value="{}".format(target_data[0].qualification),
                    )
                else:
                    ws.cell(
                        row=row_counter,
                        column=2,
                        value="",
                    )

                col_counter = 3

                for entry in target_data:
                    ws.cell(
                        row=row_counter,
                        column=col_counter,
                        value="{}".format(entry.subject),
                    )
                    ws.cell(
                        row=row_counter,
                        column=col_counter + 1,
                        value="{}".format(entry.grade),
                    )
                    col_counter += 2
            else:
                raise InputError(
                    False,
                    f"Key {desired_data} given not found in dictionary {student.which_grades}",
                )

        return ws

    @staticmethod
    def populate_master_header(ws):

        ws.cell(row=1, column=1, value="UCAS ID")
        ws.cell(row=1, column=2, value="Qualification")
        ws.cell(row=1, column=3, value="Issues Importing?")
        ws.cell(row=1, column=4, value="No. Subjects")
        ws.cell(row=1, column=5, value="Math Grade")
        ws.cell(row=1, column=6, value="Physics Grade")
        ws.cell(row=1, column=11, value="FM?")
        ws.cell(row=1, column=12, value="Overall Grade")
        ws.cell(row=1, column=13, value="Cycle")
        ws.cell(row=1, column=14, value="Marker")
        ws.cell(row=1, column=15, value="Batch")

        for i in range(0, 4, 2):
            ws.cell(row=1, column=7 + i, value="Grade")
            ws.cell(row=1, column=8 + i, value="Subject")

        return ws

    @staticmethod
    def update_al_string(categorised_entries, input_string, is_fm):
        if is_fm:
            # If fm => min 3 subjects
            if categorised_entries["additional_subjects"]:
                # If the list is not empty => 4 or more subjects
                return input_string + " (A*AAA)"
            else:
                # If the list is empty => 3 subjects
                return input_string + " (A*A*A)"
        else:
            if len(categorised_entries["additional_subjects"]) > 1:
                return input_string + " (A*AAA)"
            else:
                # If the list is empty => 3 subjects
                return input_string + " (A*A*A)"

    def compile_for_master(self, ws):

        ws = self.populate_master_header(ws)

        lb = 2

        for row_counter, student in zip(
            range(lb, self.num_students + lb), self.all_students
        ):
            # Fill in UCAS ID
            ws.cell(row=row_counter, column=1, value="{}".format(student.unique_id))

            # Fill with admin details
            ws.cell(row=row_counter, column=13, value=settings.cycle)
            # What are the brackets in self.student_to_marker_mapping
            # [student.unique_id] => get value using that key
            # [0] => first value in list (value of key is a list)
            # [:2] => slice first two letters in string (initials of maker)
            ws.cell(
                row=row_counter,
                column=14,
                value=self.student_to_marker_mapping[student.unique_id][0][:2],
            )
            ws.cell(row=row_counter, column=15, value=settings.batch_number)

            # Categorise each entry into subjects
            categorised_entries = self.sort_into_subjects(student)

            # Identify if FM is presetn
            if categorised_entries["fm"]:
                is_fm = True
                ws.cell(row=row_counter, column=11, value="Yes")
            else:
                is_fm = False
                ws.cell(row=row_counter, column=11, value="No")

            # Identify and populate cell with the main qualification
            main_qualification = student.get_main_qualification()
            sanitised_string = self.internal_mapping.get(main_qualification)

            if main_qualification == "" or sanitised_string is None:
                ws.cell(row=row_counter, column=2, value="")
                ws.cell(
                    row=row_counter,
                    column=3,
                    value="Need manual entry: no valid qual found",
                )
                continue

            if "United Kingdom" in sanitised_string:
                uk_based = True
            else:
                uk_based = False

            if "A Levels" in sanitised_string:
                sanitised_string = self.update_al_string(
                    categorised_entries, sanitised_string, is_fm
                )

            # Fill in the qualification to the worksheet
            ws.cell(
                row=row_counter, column=2, value="{}".format(sanitised_string.strip())
            )

            # Create log for issues
            any_issues = self.log_issues(categorised_entries, uk_based, ws, row_counter)

            # Get all unique qualifications
            qualification = student.unique_qualifications()

            # Identify if the qualification has an overall score
            # Intersection of qualification set and set of qualifications with overall score is not empty
            intersection_qualification = (
                qualification & qualifications_with_overall_score()
            )
            # If it does, find the overall score
            if intersection_qualification:

                overall_grade = self.determine_overall_grade(
                    intersection_qualification, student
                )

                # If an overall grade exists, populate cell
                num_overall_grade = len(overall_grade)

                # If an overall grade exists
                if num_overall_grade > 1:
                    # Log multiple grades as an issue
                    if any_issues is not None:
                        any_issues.append("Multiple overall score. 1st selected.")
                    else:
                        # If any_issue is None => M&P grade not found. Skip rest of loop
                        ws.cell(
                            row=row_counter,
                            column=3,
                            value="M&P missing, need manual entry. Multiple overall score. 1st selected.",
                        )
                        output_overall_grade = self.strip_overall_grade_spaces(
                            overall_grade
                        )
                        ws.cell(
                            row=row_counter,
                            column=12,
                            value="{}".format(output_overall_grade),
                        )
                        continue

                # Populate if list is not empty
                if num_overall_grade != 0:
                    output_overall_grade = self.strip_overall_grade_spaces(
                        overall_grade
                    )
                    ws.cell(
                        row=row_counter,
                        column=12,
                        value="{}".format(output_overall_grade),
                    )

            # If M&P missing, clearly there is an issue. Skip the rest of loop
            if any_issues is None:
                ws.cell(
                    row=row_counter, column=3, value="M&P missing, need manual entry"
                )
                continue

            self.populate_grades(
                categorised_entries,
                ws,
                is_fm,
                row_counter,
                any_issues,
            )

            # Compress list of strings into a single string
            any_issues = self.compress_log(any_issues)
            # Populate cell with string
            if any_issues is None:
                ws.cell(row=row_counter, column=3, value="")
            else:
                ws.cell(row=row_counter, column=3, value=any_issues)

        return ws

    @staticmethod
    def strip_overall_grade_spaces(overall_grades):
        if len(overall_grades) > 0:
            output_overall_grade = overall_grades[0]
        else:
            output_overall_grade = overall_grades

        if isinstance(output_overall_grade, str):
            output_overall_grade = output_overall_grade.strip()
        else:
            output_overall_grade = str(output_overall_grade).strip()

        return output_overall_grade

    @staticmethod
    def sanitise_grade_of_pass(grade_val):

        if grade_val is None:
            # If grade is None, replace with a dash
            return "-"
        elif not isinstance(grade_val, str):
            grade_val = str(grade_val).strip()

        if "(pass)" in grade_val:
            return grade_val.replace("(pass)", "").strip()
        elif "(Pass)" in grade_val:
            return grade_val.replace("(Pass)", "").strip()
        elif "pass" in grade_val:
            return grade_val.replace("pass", "").strip()
        elif "Pass" in grade_val:
            return grade_val.replace("Pass", "").strip()
        else:
            return grade_val.strip()

    def populate_grades(
        self,
        categorised_entries,
        ws,
        is_fm,
        row_counter,
        any_issues,
    ):
        # Populate subject and grades
        subject_counter = 0
        map_subject_num_to_cols = {0: [5], 1: [6], 2: [7, 8], 3: [9, 10]}

        for subject_entries in categorised_entries.values():

            # If no entries, then go to next value
            if not subject_entries:
                if subject_counter > 1:
                    continue
                else:
                    subject_counter += 1
                    continue

            excel_col = map_subject_num_to_cols.get(subject_counter)
            # If there is only one subject entry, then just populate
            if len(subject_entries) == 1:
                for col, val in zip(excel_col, subject_entries[0].grade_info):
                    val = self.sanitise_grade_of_pass(val)
                    ws.cell(row=row_counter, column=col, value=val)

            # if it is at the 3rd subject, and not FM then iterate over all
            elif subject_counter == 2 and not is_fm:
                for entry in subject_entries[:2]:
                    excel_col = map_subject_num_to_cols.get(subject_counter)
                    for col, val in zip(excel_col, entry.grade_info):
                        val = self.sanitise_grade_of_pass(val)
                        ws.cell(row=row_counter, column=col, value=val)
                    subject_counter += 1

            # If there is FM and there are more than 4 subjects, populate with 1st
            elif subject_counter == 3 and len(subject_entries) > 1:
                for col, val in zip(excel_col, subject_entries[0].grade_info):
                    val = self.sanitise_grade_of_pass(val)
                    ws.cell(row=row_counter, column=col, value=val)

            # Doesn't fall into any of above cases => special case
            else:

                exam_result_entries = [
                    entry for entry in subject_entries if entry.is_exam_result
                ]
                predicted_entries = [
                    entry for entry in subject_entries if entry.is_predicted
                ]
                excel_col = excel_col

                map_counter_to_subject = {0: "math", 1: "physics", 2: "3rd", 3: "4th"}
                any_issues.append(
                    "For {}, selected ".format(map_counter_to_subject[subject_counter])
                )

                # # REFACTOR
                if len(exam_result_entries) == 1:
                    any_issues[-1] += "exam result"
                    selected = exam_result_entries[0]
                elif len(exam_result_entries) > 1:
                    any_issues[-1] += "exam result"
                    selected = self.select_an_entry(exam_result_entries, any_issues)
                elif len(predicted_entries) == 1:
                    any_issues[-1] += "predicted grade"
                    selected = predicted_entries[0]
                elif len(predicted_entries) > 1:
                    any_issues[-1] += "predicted grade"
                    selected = self.select_an_entry(predicted_entries, any_issues)
                else:
                    any_issues[-1] += "from all"
                    selected = self.select_an_entry(subject_entries, any_issues)

                for col, val in zip(excel_col, selected.grade_info):
                    val = self.sanitise_grade_of_pass(val)
                    ws.cell(row=row_counter, column=col, value=val)

            subject_counter += 1

        overall_grade = ws.cell(row=row_counter, column=12).value
        if overall_grade is None:
            self.populate_alphabetic_overall_grade(ws, row_counter)

    @staticmethod
    def populate_alphabetic_overall_grade(ws, row_counter):
        overall_grade = ""
        target_cols = [5, 6, 7, 9]

        for col in target_cols:
            grade_val = ws.cell(row=row_counter, column=col).value
            if grade_val is None:
                if col != 9:
                    # If final one is empty, don't append anything
                    # If it isn't the final subject, put a _ to indicate missing
                    overall_grade += "_"
            else:
                grade_val = str(grade_val).strip()

                if re.search(r"[a-zA-z*-]", grade_val) is None:
                    # If any of the values are numeric, exit function
                    return
                else:
                    overall_grade += grade_val

        ws.cell(row=row_counter, column=12, value="{}".format(overall_grade))

    @staticmethod
    def determine_overall_grade(intersection_qualification, student):
        # Is it IB?
        if intersection_qualification & ib_permutations():
            # It is IB as intersection of qualification and IB perms is not empty
            if student.results_entries:
                overall_grade = [
                    item.grade
                    for item in student.results_entries
                    if "IB Total points" in item.qualification
                ]
            elif student.completed_entries:
                overall_grade = [
                    item.grade
                    for item in student.completed_entries
                    if "International Baccalaureate Diploma" in item.qualification
                ]
            elif student.predicted_entries:
                overall_grade = [
                    item.grade
                    for item in student.predicted_entries
                    if "International Baccalaureate Diploma" in item.qualification
                ]
            else:
                overall_grade = []
        else:
            intersection_qualification = list(intersection_qualification)
            # Get all grades that aren't none
            overall_grade = [
                item
                for qualification in intersection_qualification
                for item in student.get_grade_for_qualification(qualification)
            ]

        return overall_grade

    @staticmethod
    def select_an_entry(lst_of_entries, any_issues):
        # ASSUMPTION: MOST RECENT GRADE = BEST GRADE

        years = [entry.year for entry in lst_of_entries if entry.year is not None]

        # If the list is empty then look for highest grade
        if not years:
            grades = [
                str(entry.grade) for entry in lst_of_entries if entry.grade is not None
            ]
            # If this list is not empty, no issues. Take the highest
            if grades:
                any_issues[-1] += "highest "
                return lst_of_entries[grades.index(max(grades))]
            else:
                any_issues[-1] = "No grade"
                return GradeEntry(None, None, "-", False, None, False)

        # Determine if all are unique and whether the lsit is empty
        if len(years) == len(set(years)) and years:
            # All unique years
            any_issues[-1] += "most recent "

            return lst_of_entries[years.index(max(years))]

        max_year = max(years)
        most_recent = [entry for entry in lst_of_entries if entry.year == max_year]
        if len(most_recent) > 1:
            # Max year is repeated

            grades = [
                str(entry.grade) for entry in most_recent if entry.grade is not None
            ]
            any_issues[-1] += "highest most recent "

            return lst_of_entries[grades.index(max(grades))]

    @staticmethod
    def compress_log(log):
        if len(log) > 1:
            return " ".join(log)
        elif len(log) == 1:
            return log[0]
        else:
            return None

    @staticmethod
    def log_issues(categorised_entries, uk_based, ws, rowCounter):
        convert_lst_to_bool = [
            not bool(cat_entry) for cat_entry in categorised_entries.values()
        ]
        # All lists in categorised entries are empty
        if all(convert_lst_to_bool):
            return "No valid qual & subjects."

        if all(convert_lst_to_bool[:3]):
            return None

        log = []

        for subject, entries in categorised_entries.items():
            entries_length = len(entries)
            if entries_length > 1 and subject != "additional_subjects":
                log.append("Multiple {} qual.".format(subject))
            elif entries_length > 2 and subject == "additional_subjects":
                if not uk_based:
                    # log.append("Multiple Subjects.")
                    ws.cell(row=rowCounter, column=4, value="Multiple")
                else:
                    # log.append(">4 Subjects.")
                    ws.cell(row=rowCounter, column=4, value=">4")
            elif (
                entries_length == 0
                and subject != "fm"
                and subject != "additional_subjects"
            ):
                log.append("{} missing.".format(subject))

        return log

    @staticmethod
    def sort_into_subjects(student):

        categorised_entries = {
            "math": [],
            "physics": [],
            "fm": [],
            "additional_subjects": [],
        }

        # Grade type => results/predicted/completed
        for grade_entries in student.which_grades.values():
            # List of entries are not empty
            if grade_entries:
                for entry in grade_entries:
                    if entry.subject in math_mapping().get(entry.qualification, set()):
                        categorised_entries["math"].append(entry)
                    elif entry.subject in physics_mapping().get(
                        entry.qualification, set()
                    ):
                        categorised_entries["physics"].append(entry)
                    elif entry.subject in fm_mapping().get(entry.qualification, set()):
                        categorised_entries["fm"].append(entry)
                    else:
                        categorised_entries["additional_subjects"].append(entry)

        return categorised_entries

    def write_to_excel(self, output_abs_path):

        is_abs_path(output_abs_path)

        from openpyxl import Workbook

        wb = Workbook()

        # Create and populate workshets
        completed = wb.create_sheet("Completed Qualifications")
        completed = self.populate_worksheet("completed", completed)

        predicted = wb.create_sheet("Predicted Grades")
        predicted = self.populate_worksheet("predicted", predicted)

        exam_results = wb.create_sheet("Exam Results")
        exam_results = self.populate_worksheet("results", exam_results)

        compiled_single = wb.create_sheet("Compiled", 0)
        compiled_single = self.compile_for_master(compiled_single)

        wb.save(os.path.join(output_abs_path, settings.output_filename))
